from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
import db.crud.admins as crud_admins
from bot.default_functions import send_callback_aiogram_message, send_message_aiogram_message

import bot.admin.texts as texts
import bot.admin.keyboards as keyboards
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re


def is_admin(func):
    async def wrapper(message: Message | CallbackQuery, state: FSMContext):
        peer_id = int(message.from_user.id)
        if await crud_admins.read_admin(peer_id):
            return await func(message, state)
        else:
            if isinstance(message, Message):
                return await send_message_aiogram_message(
                    message=message,
                    text=texts.admin_no_access,
                    keyboard=await keyboards.to_user_menu()
                )
            return await send_callback_aiogram_message(
                callback=message,
                text=texts.admin_no_access,
                keyboard=await keyboards.to_user_menu()
            )

    return wrapper


def decline_track(number):
    if not isinstance(number, int) or number < 0:
        return f"{number} треков"

    last_two = number % 100
    if 11 <= last_two <= 19:
        return f"{number} треков"

    last_one = number % 10
    if last_one == 1:
        return f"{number} трек"
    elif 2 <= last_one <= 4:
        return f"{number} трека"
    else:
        return f"{number} треков"


def export_users_to_excel_pandas(users, filename: str = "users_export.xlsx"):
    data = [
        {
            "ID": i,
            "Айди": users[i].peer_id,
            "Ник": users[i].username,
            "Кол-во треков": users[i].count,
            "Предпочтения": users[i].preference
        }
        for i in range(len(users))
    ]
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    for col in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col)
        max_length = max(
            df.iloc[:, col - 1].astype(str).map(len).max(),
            len(str(df.columns[col - 1])))
        ws.column_dimensions[column_letter].width = max_length + 2
        header_cell = ws[f"{column_letter}1"]
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
    wb.save(filename)
    return filename


def check_playlist_link(playlist_url):
    link = re.search(r"playlists/(\d+)", playlist_url)
    if link is None:
        return False
    return link.group(1)
